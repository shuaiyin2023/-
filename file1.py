# import matplotlib.pyplot as plt
#
#
# def statistics_age(age_data):
#     age_ranges = [(1, 10), (11, 20), (21, 30), (31, 40), (41, 50), (51, 60), (61, 70), (71, 80), (81, 90), (91, 100),
#                   (101, 110), (111, 120)]
#     age_range_dict = {f"{start}-{end}": 0 for start, end in age_ranges}
#     for age in age_data:
#         for age_range in age_ranges:
#             if age_range[0] <= age <= age_range[1]:
#                 age_range_dict[f"{age_range[0]}-{age_range[1]}"] += 1
#                 break
#         else:
#             age_range_dict["111-120"] += 1
#     return age_range_dict
#
#
# input_data = input("请输入一组年龄数据: ")
# input_data_list = [int(x) for x in input_data.split(",") if int(x) >= 1 and int(x) <= 120]
# result = statistics_age(input_data_list)
# print("统计结果如下: \n", result)
#
# # 绘制饼图
# # 设置中文字体
# plt.rcParams['font.sans-serif'] = ['SimHei']  # 使用 "SimHei" 字体
# plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号
# # 提取年龄段和对应的人数
# age_ranges = list(result.keys())
# population = list(result.values())
#
# # 绘制饼图
# plt.figure(figsize=(8, 8))
# plt.pie(population, labels=age_ranges, autopct='%1.1f%%', startangle=140)
# plt.axis('equal')  # 使饼图呈圆形
# plt.title("不同年龄段的人数占比")  # 添加标题
#
# # 显示饼图
# plt.show()
#
import time

# def reward_student(score):
#     if score == 100:
#         return "奖励一辆BMW"
#     elif score >= 80:
#         return "奖励一台iphone"
#     elif score >= 60:
#         return "奖励一本参考书"
#     else:
#         return "再接再厉哦"
#
# score_data = input("请输入五个学生成绩: ").split(",")
# for score in score_data:
#     score = int(score)
#     print(reward_student(score))


# def record_time(func):
#     def wrapper(*arg, **kwargs):
#         start = time.time()
#         result = func(*arg, **kwargs)
#         end = time.time()
#         print("总耗时: ", end - start)
#         return result
#
#     return wrapper
#
#
# @record_time
# def download(filename):
#     print(f"开始下载文件--{filename}")
#     time.sleep(2)
#     print(f"文件--{filename}下载完成")
#
#
# @record_time
# def upload(filename):
#     print(f"开始上传文件--{filename}")
#     time.sleep(2)
#     print(f"文件--{filename}上传完成")
#
#
# file1 = "<<python入门>>"
# print(download(file1))
# print(upload(file1))


# def fac(num):
#     if num in (0, 1):
#         return 1
#     return num * fac(num - 1)
#
#
# print(fac(5))
#
#
# def fib(n):
#     if n in (1,2):
#         return 1
#     return fib(n-1) + fib(n-2)
# for i in range(1, 10):
#     print(fib(i))
#
#
# def fib(n):
#     a, b = 0, 1
#     for _ in range(n):
#         a, b = b, a + b
#     return a
#
# print(fib(5))


import time

import rqdatac


# 定义数字时钟类
# class Clock(object):
#     """数字时钟"""
#
#     def __init__(self, hour=0, minute=0, second=0):
#         """初始化方法
#         :param hour: 时
#         :param minute: 分
#         :param second: 秒
#         """
#         self.hour = hour
#         self.min = minute
#         self.sec = second
#
#     def run(self):
#         """走字"""
#         self.sec += 1
#         if self.sec == 60:
#             self.sec = 0
#             self.min += 1
#             if self.min == 60:
#                 self.min = 0
#                 self.hour += 1
#                 if self.hour == 24:
#                     self.hour = 0
#
#     def show(self):
#         """显示时间"""
#         return f'{self.hour:0>2d}:{self.min:0>2d}:{self.sec:0>2d}'
#
#
# # 创建时钟对象
# clock = Clock(23, 59, 58)
# while True:
#     # 给时钟对象发消息读取时间
#     print(clock.show())
#     # 休眠1秒钟
#     time.sleep(1)
#     # 给时钟对象发消息使其走字
#     clock.run()


class Person:

    # __slots__ = ('name', 'age')
    def __init__(self, name, age):
        self.__name = name
        self.age = age

    def print_info(self):
        print(f'{self.__name} is {self.age} years')

    @property
    def name(self):
        return self.__name

    @staticmethod
    def userinfo(name, age, gender):
        return f"{name}的age是{age},并且是一名{gender}"

    def eat(self):
        print(f"{self.__name}正在吃饭")


class Student(Person):
    def __init__(self, name, age):
        super().__init__(name, age)

    def study(self):
        print(f"{self.name}正在学习")


class Teacher(Person):
    def __init__(self, name, age, title):
        super().__init__(name, age)
        self.title = title

    def teach(self, course_name):
        print(f"{self.name}今年{self.age}，职称是{self.title}，正在讲课{course_name}")

    def eat(self):
        return f"{self.name}肚子饿了，想吃饭"


# p1 = Person('yin', 22)
# print(p1.print_info())
# p1.gender = "male"
# print(p1.gender)
# print("===========================分割线===============================")
# print(p1.age)
# print(p1._Person__name)
# print(p1.name)

# print(p1.userinfo("yin", 22, "male"))
# print("=======================分割线=====================")
# stu1 = Student('yin', 22)
# print(stu1.eat())
#
# tea1 = Teacher('yin', 22, "教授")
# print(tea1.teach("python程序入门"))
#
# tea2 = Teacher('yin', 22, "副教授")
# print(tea2.eat())
#
#
#
# import base64
#
# str1 = "my name is liangzai"
# base64_str = base64.b64encode(str1.encode())
# print(base64_str)
# base64_s = base64.b64decode(base64_str).decode()
# print(base64_s)


# from collections import Counter
#
# words = [
#     'look', 'into', 'my', 'eyes', 'look', 'into', 'my', 'eyes',
#     'the', 'eyes', 'the', 'eyes', 'the', 'eyes', 'not', 'around',
#     'the', 'eyes', "don't", 'look', 'around', 'the', 'eyes',
#     'look', 'into', 'my', 'eyes', "you're", 'under'
# ]
# counter = Counter(words)
# # 打印words列表中出现频率最高的3个元素及其出现次数
# for elem, count in counter.most_common(3):
#     print(elem, count)
#
#
# import hashlib
#
# # 计算字符串"123456"的MD5摘要
# print(hashlib.md5('123456'.encode()).hexdigest())
# hashlib.md5('123456'.encode()).hexdigest()
# print(hashlib.md5('mumu962464'.encode()).hexdigest())
#
# # 计算文件"Python-3.7.1.tar.xz"的MD5摘要
# hasher = hashlib.md5()
# with open('D:\Company-yin\Files\新建文本文档.txt', 'rb') as file:
#     data = file.read(512)
#     while data:
#         hasher.update(data)
#         data = file.read(512)
# print(hasher.hexdigest())
#
#
# import itertools
#
# # 产生ABCD的全排列
# for value in itertools.permutations('ABCD'):
#     print(value)
#
#
# import uuid
# print(uuid.uuid1())
# print(uuid.uuid1().hex)


# import rqdatac as rq
#
# rq.init(uri='tcp://13774400687:021700@10.66.200.116:16010')
#
#
# result = rq.get_trading_dates("20240401", "20240430", market='cn')
# print(result)
# df = rq.futures.get_dominant_price(underlying_symbols="IC", start_date="20210104", end_date="20210105", frequency='1m',
#                                    adjust_type='none', adjust_method='prev_close_spread')
# print(df)


import random

import openpyxl

# 第一步：创建工作簿（Workbook）
# wb = openpyxl.Workbook()
#
# # 第二步：添加工作表（Worksheet）
# sheet = wb.active
# sheet.title = '期末成绩'
#
# titles = ('姓名', '语文', '数学', '英语')
# for col_index, title in enumerate(titles):
#     sheet.cell(1, col_index + 1, title)
#
# names = ('关羽', '张飞', '赵云', '马超', '黄忠')
# for row_index, name in enumerate(names):
#     sheet.cell(row_index + 2, 1, name)
#     for col_index in range(2, 5):
#         sheet.cell(row_index + 2, col_index, random.randrange(50, 101))

# 第四步：保存工作簿
# wb.save('考试成绩表.xlsx')


# from openpyxl import Workbook
# from openpyxl.chart import BarChart, Reference
#
# wb = Workbook(write_only=True)
# sheet = wb.create_sheet()
#
# rows = [
#     ('类别', '销售A组', '销售B组'),
#     ('手机', 40, 30),
#     ('平板', 50, 60),
#     ('笔记本', 80, 70),
#     ('外围设备', 20, 10),
# ]
#
# # 向表单中添加行
# for row in rows:
#     sheet.append(row)
#
# # 创建图表对象
# chart = BarChart()
# chart.type = 'col'
# chart.style = 10
# # 设置图表的标题
# chart.title = '销售统计图'
# # 设置图表纵轴的标题
# chart.y_axis.title = '销量'
# # 设置图表横轴的标题
# chart.x_axis.title = '商品类别'
# # 设置数据的范围
# data = Reference(sheet, min_col=2, min_row=1, max_row=5, max_col=3)
# # 设置分类的范围
# cats = Reference(sheet, min_col=1, min_row=2, max_row=5)
# # 给图表添加数据
# chart.add_data(data, titles_from_data=True)
# # 给图表设置分类
# chart.set_categories(cats)
# chart.shape = 4
# # 将图表添加到表单指定的单元格中
# sheet.add_chart(chart, 'A10')
#
# wb.save('demo.xlsx')


from docx import Document
from docx.document import Document as Doc

# 将真实信息用字典的方式保存在列表中
# employees = [
#     {
#         'name': '骆昊',
#         'id': '100200198011280001',
#         'sdate': '2008年3月1日',
#         'edate': '2012年2月29日',
#         'department': '产品研发',
#         'position': '架构师',
#         'company': '成都华为技术有限公司'
#     },
#     {
#         'name': '王大锤',
#         'id': '510210199012125566',
#         'sdate': '2019年1月1日',
#         'edate': '2021年4月30日',
#         'department': '产品研发',
#         'position': 'Python开发工程师',
#         'company': '成都谷道科技有限公司'
#     },
#     {
#         'name': '李元芳',
#         'id': '2102101995103221599',
#         'sdate': '2020年5月10日',
#         'edate': '2021年3月5日',
#         'department': '产品研发',
#         'position': 'Java开发工程师',
#         'company': '同城企业管理集团有限公司'
#     },
# ]
# # 对列表进行循环遍历，批量生成Word文档
# for emp_dict in employees:
#     # 读取离职证明模板文件
#     doc = Document('resources/离职证明模板.docx')  # type: Doc
#     # 循环遍历所有段落寻找占位符
#     for p in doc.paragraphs:
#         if '{' not in p.text:
#             continue
#         # 不能直接修改段落内容，否则会丢失样式
#         # 所以需要对段落中的元素进行遍历并进行查找替换
#         for run in p.runs:
#             if '{' not in run.text:
#                 continue
#             # 将占位符换成实际内容
#             start, end = run.text.find('{'), run.text.find('}')
#             key, place_holder = run.text[start + 1:end], run.text[start:end + 1]
#             run.text = run.text.replace(place_holder, emp_dict[key])
#     # 每个人对应保存一个Word文档
#     doc.save(f'{emp_dict["name"]}离职证明.docx')


# import re
#
# sentence = 'Oh, shit! 你是傻逼吗? Fuck you.'
# purified = re.sub('fuck|shit|[傻煞沙][比笔逼叉缺吊碉雕]',
#                   '*', sentence, flags=re.IGNORECASE)
# print(purified)  # Oh, *! 你是*吗? * you.
#
#
# import re
#
# poem = '窗前明月光，疑是地上霜。举头望明月，低头思故乡。'
# sentences_list = re.split(r'[，。]', poem)
# sentences_list = [sentence for sentence in sentences_list if sentence]
# for sentence in sentences_list:
#     print(sentence)


import copy

print("====================赋值=====================")
a = [1, 2, [3, 4]]
b = a

print(a, id(a))
print(b, id(b))

b = 2
print(a, id(a))
print(b, id(b))

print("====================浅拷贝=====================")
c = [1, 2, [3, 4]]
d = copy.copy(c)
print(c, id(c))
print(d, id(d))

d = 2
print(c, id(c))
print(d, id(d))

print("====================深拷贝=====================")
e = [1, 2, [3, 4]]
f = copy.deepcopy(e)
print(e, id(e))
print(f, id(f))

f = 2
print(e, id(e))
print(f, id(f))

# def trim(s):
#     if s == '':
#         return s
#     if s[0] == ' ':
#         return trim(s[1:])
#     if s[-1] == ' ':
#         return trim(s[:-1])
#
#     return s
#
# if trim('hello  ') != 'hello':
#     print(1)
#     print('测试失败!')
# elif trim('  hello') != 'hello':
#     print(2)
#     print('测试失败!')
# elif trim('  hello  ') != 'hello':
#     print(3)
#     print('测试失败!')
# elif trim('  hello  world  ') != 'hello  world':
#     print(4)
#     print('测试失败!')
# elif trim('') != '':
#     print(5)
#     print('测试失败!')
# elif trim('    ') != '':
#     print(6)
#     print('测试失败!')
# else:
#     print('测试成功!')


# def findMinAndMax(L):
#     if len(L) == 0:
#         return (None, None)
#
#     min_num = max_num = L[0]
#     for i in range(1, len(L)):
#         if L[i] < min_num:
#             min_num = L[i]
#         if L[i] > max_num:
#             max_num = L[i]
#     return (min_num, max_num)


# if findMinAndMax([]) != (None, None):
#     print('测试失败!')
# elif findMinAndMax([7]) != (7, 7):
#     print('测试失败!')
# elif findMinAndMax([7, 1]) != (1, 7):
#     print('测试失败!')
# elif findMinAndMax([7, 1, 3, 9, 5]) != (1, 9):
#     print('测试失败!')
# else:
#     print('测试成功!')


L1 = ["World", "Hello", 18, None]
L2 = [lis.lower() for lis in L1 if isinstance(lis, str)]
print(L2)

a, b = 1, 2
print(a, b)
a, b = b, a + b
print(a, b)

'''    杨辉三角 
          1
         / \
        1   1
       / \ / \
      1   2   1
     / \ / \ / \
    1   3   3   1
   / \ / \ / \ / \
  1   4   6   4   1
 / \ / \ / \ / \ / \
1   5   10  10  5   1
'''


# 左右两端和顶端的边界值都是1
# 第n行有n个值
# 每行数字左右对称


# n行数据
# n = 10
# pascal_triangle = [[1], [1, 1]]
# for i in range(2, n):  # 按行生成数据
#     row_list = list()  # 子行
#     row_list.append(1)
#
#     for j in range(1, i):
#         row_list.append(pascal_triangle[i - 1][j - 1] + pascal_triangle[i - 1][j])
#     row_list.append(1)
#     pascal_triangle.append(row_list)

# n = 10
# pascal_triangle = [[1], [1, 1]]
# for i in range(2, n):  # 按行生成数据
#     row_list = [1] + [0] * (i - 1) + [1]  # 子行
#     for j in range(1, i):
#         row_list[j] = pascal_triangle[i - 1][j - 1] + pascal_triangle[i - 1][j]
#     pascal_triangle.append(row_list)

def generate_pascal_triangle(n):
    pascal_triangle = [[1] * (i + 1) for i in range(n)]
    for i in range(2, n):
        for j in range(1, i):
            pascal_triangle[i][j] = pascal_triangle[i - 1][j - 1] + pascal_triangle[i - 1][j]
    return pascal_triangle


n = 10
pascal_triangle = generate_pascal_triangle(n)

print(f"行数为{n}的杨辉三角结果: \n", pascal_triangle)

for row in pascal_triangle:
    print(row)


def normalize(name):
    name = name[0].upper() + name[1:].lower()
    return name


# L1 = ['adam', 'LISA', 'barT']
# L2 = list(map(normalize, L1))
# print(L2)
#
#
# from functools import reduce
# def prod(L):
#     def compute(x,y):
#         print(x, y)
#         return x * y
#     return reduce(compute, L)

# print('3 * 5 * 7 * 9 =', prod([3, 5, 7, 9]))
# if prod([3, 5, 7, 9]) == 945:
#     print('测试成功!')
# else:
#     print('测试失败!')
#
# import rqsdk, rqdatac, rqdatac_fund
# rqdatac.init(uri='tcp://13774400687:021700@10.66.200.116:16010')
# res = rqdatac.get_trading_dates(20240411, 20240430, market='cn')
# res = rqdatac.get_price("IC2303", start_date='2013-01-04', end_date='2014-01-04', frequency='1d', fields=None, adjust_type='pre', skip_suspended =False, market='cn', expect_df=True,time_slice=None)
# res = rqdatac.all_instruments(type=None, market='cn', date=None)

# print(res)

# r = rqdatac.all_instruments(type=None, date='2013-01-04', market='cn')
# print(r)


class Goods:
    """ this is a class """

    def __init__(self):
        self.price = 100
        self.discount = 0.8

    # def __str__(self):
    #     return f"price is: {self.price}, discount is: {self.discount}"

    def get_price(self):
        return self.price * self.discount

    def set_price(self, value):
        self.price = value

    def del_price(self):
        """ hello """
        del self.price

    PRICE = property(get_price, set_price, del_price, "价格描述")


goods = Goods()
print(goods.PRICE)
goods.PRICE = 20
print(goods.PRICE)
print(goods.PRICE.__doc__)
print(goods.__doc__)
print(goods.del_price.__doc__)
print(goods.__class__)
print(goods.__module__)

print(goods.__dict__)
good_item = goods.__dict__
good_item["price"] = 100
print(good_item)
print(goods.__dict__)


class Student(Goods):

    def __init__(self, name, age):
        self.name = name
        self.age = age

    def __str__(self):
        return f'{self.name} is {self.age} years'


print(Student.__bases__)

object = Student("yinshuai", 23)
print(str(object))



import logging


logging.basicConfig(
    level=logging.INFO,
    filename="log.txt",
    filemode="w",
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s - %(funcName)s"
)

logging.info("this is info level")
logging.debug("this is debug level")
logging.warning("this is warning level")
logging.error("this is error level")
logging.critical("this is critical level")

def log():
    for i in range(10):
        if i % 2 != 0:
            logging.info(f"{i} can not be divided by 2")
    return 0
print(log())



import rqsdk, rqdatac, rqdatac_fund
rqdatac.init(uri='tcp://13774400687:021700@10.66.200.116:16010')
res = rqdatac.get_trading_dates(20240501, 20240531, market='cn')
print(res)



for i in range(1, 10):
    for j in range(1, i + 1):
        print(f"{j}*{i}={i * j}   ", end='')
    print("\n")